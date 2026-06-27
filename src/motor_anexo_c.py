# gerar_anexo_c.py

import re
import fitz  # pip install pymupdf
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


PDF_ENTRADA = "relatorio_fotografico_s5.pdf"
XLSX_SAIDA = "anexo_c_auto_preenchido.xlsx"

QTD_TRELICAS = 18
QTD_TERCAS = 18


CORES_STATUS = {
    "APROVADO": "00B050",
    "PINTURA": "FFFF00",
    "REPARO": "FFC000",
    "REPROVADO": "FF0000",
}

PESO_STATUS = {
    "APROVADO": 1,
    "PINTURA": 2,
    "REPARO": 3,
    "REPROVADO": 4,
}


def extrair_texto_pdf(caminho_pdf):
    doc = fitz.open(caminho_pdf)
    texto = ""

    for pagina in doc:
        texto += pagina.get_text() + "\n"

    return texto


def extrair_registros(texto):
    padrao = re.compile(
        r"(RF-\d+)\s*-\s*(.*?)\s*STATUS:\s*(APROVADO|PINTURA|REPARO|REPROVADO)",
        re.DOTALL
    )

    registros = []

    for match in padrao.finditer(texto):
        rf = match.group(1).strip()
        legenda = " ".join(match.group(2).split())
        status = match.group(3).strip()

        registros.append({
            "rf": rf,
            "legenda": legenda,
            "status": status,
        })

    return registros


def pior_status(status_atual, novo_status):
    if not status_atual:
        return novo_status

    if PESO_STATUS[novo_status] > PESO_STATUS[status_atual]:
        return novo_status

    return status_atual


def numeros_da_terca(legenda):
    m = re.search(r"Terças?\s+(\d{1,2})(?:\s+e\s+(\d{1,2}))?", legenda, re.I)

    if not m:
        return []

    numeros = [int(m.group(1))]

    if m.group(2):
        numeros.append(int(m.group(2)))

    return numeros


def numeros_da_trelica(legenda):
    m = re.search(r"Vão entre Treliças\s+(\d{1,2})\s+e\s+(\d{1,2})", legenda, re.I)

    if m:
        return [int(m.group(1)), int(m.group(2))]

    m = re.search(r"Treliça\s+(\d{1,2})", legenda, re.I)

    if m:
        return [int(m.group(1))]

    return []


def lado_da_foto(legenda):
    if "Lado Direito" in legenda:
        return "DIR"

    if "Lado Esquerdo" in legenda:
        return "ESQ"

    return "GERAL"


def especificidade(tercas, trelicas, legenda):
    if tercas and trelicas:
        return 3

    if "Suporte" in legenda and trelicas:
        return 2

    if tercas and not trelicas:
        return 1

    return 0


def criar_anexo_c(registros):
    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo C"

    ws.merge_cells("A1:S1")
    ws["A1"] = "ANEXO C – MAPA DE CRITICIDADE DAS TRELIÇAS E TERÇAS"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Obra/Local:"
    ws["A4"] = "Data da Inspeção:"
    ws["A5"] = "Responsável:"

    coluna_inicio = 2
    linha_inicio = 8

    ws["A8"] = "TERÇAS"

    for t in range(1, QTD_TRELICAS + 1):
        ws.cell(row=8, column=coluna_inicio + t - 1).value = f"T{t}"

    linhas_tercas = {}
    linha = 9

    for lado in ["ESQ", "DIR"]:
        for p in range(QTD_TERCAS, 0, -1):
            ws.cell(row=linha, column=1).value = f"P{p:02d} {lado}"
            linhas_tercas[(p, lado)] = linha
            linha += 1

    mapa = {}

    def marcar(p, lado, t, status, nivel, rf, legenda):
        if p < 1 or p > QTD_TERCAS:
            return

        if t < 1 or t > QTD_TRELICAS:
            return

        chave = (p, lado, t)
        atual = mapa.get(chave)

        if not atual:
            mapa[chave] = {
                "status": status,
                "nivel": nivel,
                "rf": rf,
                "legenda": legenda,
            }
            return

        if nivel > atual["nivel"]:
            mapa[chave] = {
                "status": status,
                "nivel": nivel,
                "rf": rf,
                "legenda": legenda,
            }
            return

        if nivel == atual["nivel"]:
            pior = pior_status(atual["status"], status)
            mapa[chave]["status"] = pior
            mapa[chave]["rf"] = rf
            mapa[chave]["legenda"] = legenda

    for r in registros:
        rf = r["rf"]
        legenda = r["legenda"]
        status = r["status"]

        tercas = numeros_da_terca(legenda)
        trelicas = numeros_da_trelica(legenda)
        lado = lado_da_foto(legenda)
        lados = ["DIR", "ESQ"] if lado == "GERAL" else [lado]

        nivel = especificidade(tercas, trelicas, legenda)

        # 1. Terça + trecho específico entre treliças
        if tercas and trelicas:
            for p in tercas:
                for t in trelicas:
                    for l in lados:
                        marcar(p, l, t, status, 3, rf, legenda)

        # 2. Terça sem treliça: marca a terça inteira
        elif tercas and not trelicas:
            for p in tercas:
                for t in range(1, QTD_TRELICAS + 1):
                    for l in lados:
                        marcar(p, l, t, status, 1, rf, legenda)

        # 3. Suporte lado direito: altura da terça P04
        elif "Suporte" in legenda and lado == "DIR":
            if trelicas:
                for t in trelicas:
                    marcar(4, "DIR", t, status, 2, rf, legenda)
            else:
                for t in [9, 10]:
                    marcar(4, "DIR", t, status, 2, rf, legenda)

        # 4. Suporte lado esquerdo: altura da terça P15
        elif "Suporte" in legenda and lado == "ESQ":
            if trelicas:
                for t in trelicas:
                    marcar(15, "ESQ", t, status, 2, rf, legenda)
            else:
                for t in [9, 10]:
                    marcar(15, "ESQ", t, status, 2, rf, legenda)

        # 5. Suporte sem lado: considerar região central T09 e T10
        elif "Suporte" in legenda and lado == "GERAL":
            for t in [9, 10]:
                for l in ["DIR", "ESQ"]:
                    marcar(9, l, t, status, 2, rf, legenda)

        # 6. Treliça sem terça e sem suporte: marca a coluna inteira
        elif trelicas and not tercas:
            for t in trelicas:
                for p in range(1, QTD_TERCAS + 1):
                    for l in lados:
                        marcar(p, l, t, status, 0, rf, legenda)

    for (p, lado), linha_excel in linhas_tercas.items():
        for t in range(1, QTD_TRELICAS + 1):
            item = mapa.get((p, lado, t))
            celula = ws.cell(row=linha_excel, column=coluna_inicio + t - 1)

            if item:
                status = item["status"]
            else:
                status = "APROVADO"

            celula.value = status
            celula.fill = PatternFill("solid", fgColor=CORES_STATUS[status])

    # Legenda
    ws["U2"] = "Legenda"
    legenda_cores = [
        ("Baixa Criticidade", "00B050"),
        ("Média Criticidade", "FFFF00"),
        ("Alta Criticidade", "FFC000"),
        ("Criticidade Elevada", "FF0000"),
    ]

    for idx, (texto, cor) in enumerate(legenda_cores, start=3):
        ws.cell(row=idx, column=21).fill = PatternFill("solid", fgColor=cor)
        ws.cell(row=idx, column=22).value = texto

    # Aba de registros extraídos
    ws2 = wb.create_sheet("Registros extraídos")
    ws2.append(["RF", "Legenda", "Status", "Terças", "Treliças", "Lado"])

    for r in registros:
        ws2.append([
            r["rf"],
            r["legenda"],
            r["status"],
            ", ".join(map(str, numeros_da_terca(r["legenda"]))),
            ", ".join(map(str, numeros_da_trelica(r["legenda"]))),
            lado_da_foto(r["legenda"]),
        ])

    # Aba de células preenchidas
    ws3 = wb.create_sheet("Células preenchidas")
    ws3.append(["Terça", "Lado", "Treliça", "Status", "Nível", "RF", "Legenda"])

    for (p, lado, t), item in sorted(mapa.items()):
        ws3.append([
            f"P{p:02d}",
            lado,
            f"T{t:02d}",
            item["status"],
            item["nivel"],
            item["rf"],
            item["legenda"],
        ])

    # Formatação
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borda
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, QTD_TRELICAS + 2):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.column_dimensions["A"].width = 14

    for col in ["U", "V"]:
        ws.column_dimensions[col].width = 22

    for row in range(1, linha + 2):
        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "B9"

    # Configuração de impressão: caber em 1 folha A4 paisagem
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Área de impressão somente do mapa
    ws.print_area = "A1:S45"

    # Margens menores
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    # Centralizar na página
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    wb.save(XLSX_SAIDA)


def main():
    texto = extrair_texto_pdf(PDF_ENTRADA)
    registros = extrair_registros(texto)

    print(f"Registros encontrados: {len(registros)}")

    criar_anexo_c(registros)

    print(f"Arquivo gerado: {XLSX_SAIDA}")


if __name__ == "__main__":
    main()

def gerar_anexo_c_pdf(pdf_entrada, xlsx_saida):
    global PDF_ENTRADA, XLSX_SAIDA

    PDF_ENTRADA = pdf_entrada
    XLSX_SAIDA = xlsx_saida

    texto = extrair_texto_pdf(PDF_ENTRADA)
    registros = extrair_registros(texto)
    criar_anexo_c(registros)